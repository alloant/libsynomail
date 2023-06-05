#!/bin/python

from getpass import getpass
from datetime import datetime
from tempfile import NamedTemporaryFile
import time
import re
from pathlib import Path

from subprocess import Popen
from sys import platform as _platform

import logging

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font


from synomail.get_mail import create_despacho
from synomail import CONFIG, EXT, INV_EXT
import synomail.connection as con

TITLES = ['type','source','No','Year','Ref','Date','Content','Dept','Name','Original','Comments'] 

def read_register_files():
    files_in_outbox = con.nas.get_file_list(f"{CONFIG['despacho']}/Outbox Despacho")
    files_in_outbox.sort(reverse = True,key = lambda file: file['name'])
    notes = {}
    
    for file in files_in_outbox:
        if file['name'][:9] in ["despacho-"] and 'osheet' in file['name']:
            notes |= get_register_from_despacho(file['display_path'])
   
    for file in files_in_outbox:
        if file['name'][:9] in ["register-"] and 'osheet' in file['name']:
            update_register(file['display_path'],notes)

    return notes

def get_register_from_despacho(reg):
    notes = {}
    reg_file = con.nas.download_file(reg)
    wb = load_workbook(reg_file)
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        values = list(ws.iter_rows(values_only=True))
        for row in values[1:]:
            if row[2] and row[1]:
                n = f"000{row[2]}"[-4:]
                key = f"{row[1]}_{n}"
                key_notes = []
                for v in values[1:]:
                    if v[1] == row[1] and v[2] == row[2]:
                        key_notes.append(dict(zip(values[0],v)))
                        if v[6]:
                            key_notes[-1]['main'] = True
                            dept = row[7]
                        else:
                            key_notes[-1]['main'] = False

            else:
                key = row[8][:-2].split('","')[1]
                key_notes = [dict(zip(values[0],row))]
                key_notes[0]['main'] = True
                dept = row[7]
            
            notes[key] = {'notes':key_notes,'Dept':dept}
                       
    return notes

def update_register(reg,notes):
    reg_file = con.nas.download_file(reg)
    wb = load_workbook(reg_file)
    
    data = []
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        
        for i,row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                columns = row
            else:
                data.append(dict(zip(columns,row)))
    
    for dt in data:
        if dt['key'] in notes:
            notes[dt['key']]['archived'] = dt['archived']
            notes[dt['key']]['sent'] = dt['sent']
            notes[dt['key']]['link'] = dt['link']
            notes[dt['key']]['name_changed'] = dt['name_changed']
            if dt['Dept'] != '' and dt['Dept'] != None:
                notes[dt['key']]['Dept'] = dt['Dept']

def archive_notes(reg_notes):
    path = f"{CONFIG['despacho']}/Outbox Despacho"
    for num,data in reg_notes.items():
        if len(data['notes']) == 0:
            continue
        
        if data['Dept'] in [None,'']:
            data['link'] = ''
            data['archived'] = False
            continue

        
        dest = f"{CONFIG['archive']}/{data['type']} in {data['Year']}"
        if data['create_folder']:
            path_link = con.nas.create_folder(dest,num)
            dest = f"{dest}/{num}"
        
        archived = False
        if 'archived' in data:
            if data['archived']:
                archived = True
        
        name_changed = 0
        if 'name_changed' in data:
            name_changed = data['name_changed']
        else:
            data['name_changed'] = 0

        data['notes'].sort(reverse = True,key = lambda file: file['main'])
        moved = False
        p_link = ''
        
        error_name = False
        for note in data['notes']:
            if note['main']:
                name = note['Name'][:-2].split('","')[1]
                p_link = note['Name'][:-2].split('","')[0].split("/")[-1]
                ext = Path(name).suffix[1:]
                ext_main = ext
                
                if not archived:
                    if name_changed in [1,3]:
                        name_main = f"{num}.{ext}"
                    elif con.nas.change_name(f"{path}/{name}",f"{num}.{ext}"):
                        name_main = f"{num}.{ext}"
                        data['name_changed'] += 1
                    else:
                        error_name = True
                break
        
        if not error_name and not archived:
            for note in data['notes']:
                name = name_main if note['main'] else note['Name'][:-2].split('","')[1]

                if con.nas.move(f"{path}/{name}",dest):
                    moved = True
            
                if not note['Original'] in ['',None]:
                    ext = Path(note['Original']).suffix[1:]
                    if name_changed in [0,1]:
                        con.nas.change_name(f"{path}/{note['Original']}",f"{num}.{ext}")
                        data['name_changed'] += 2
                    con.nas.move(f"{path}/{num}.{ext}",dest)


        if not archived:
            if moved: 
                data['archived'] = True
                logging.info(f"{name} archived")
            else:
                data['archived'] = False
     
        src = re.findall('\d+',num)
        only_num = int(src[0]) if src else num
        
        if data['create_folder']:
            data['link'] = f'=HYPERLINK("#dlink=/d/f/{path_link}", "{only_num}")'
            data['link_message'] = f'<https://nas.prome.sg:5001/d/f/{path_link}|{data["source"]} {only_num}/{data["Year"][2:]}>'
        elif ext_main in INV_EXT:
            data['link'] = f'=HYPERLINK("#dlink=/oo/r/{p_link}", "{only_num}")'
            data['link_message'] = f'<https://nas.prome.sg:5001/oo/r/{p_link}|{data["source"]} {only_num}/{data["Year"][2:]}>'
        else:
            data['link'] = f'=HYPERLINK("#dlink=/d/f/{p_link}", "{only_num}")'
            data['link_message'] = f'<https://nas.prome.sg:5001/d/f/{p_link}|{data["source"]} {only_num}/{data["Year"][2:]}>'



def upload_register(wb,name,dest):        
    con.nas.upload_convert_wb(wb,name,dest) 
    
def create_register(ws,reg_notes):
    reg_titles = ['source','link','Year','Ref','Date','Content','Dept','of_anex','archived','sent','key','name_changed']
    ws.append(reg_titles)

    font = Font(name= 'Arial',
                size=12,
                bold=False,
                italic=False,
                strike=False,
                underline='none'
                #color='4472C4'
                )

    for num,data in reg_notes.items():
        row = []
        for title in reg_titles:
            if title == 'of_anex':
                n = len(data['notes']) - 1
                row.append(n) if n > 0  else row.append('')
            elif title == 'key':
                row.append(num)
            elif title in data:
                row.append(data[title])
            else:
                row.append('-')

        #n = len(data['notes']) - 1
        #row.append(n) if n > 0  else row.append('')

        ws.append(row)
        #ws[ws.max_row][4].value = datetime.strptime(data['Date'],"%d/%m/%Y")
        ws[ws.max_row][4].number_format = 'dd/mm/yyyy;@'
        
        column_widths = [10,10,10,12,12,50,12,12]
        for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
            ws.column_dimensions[get_column_letter(i)].width = column_width
            
        for row in ws[1:ws.max_row]:  # skip the header
            for i,col in enumerate(column_widths):
                cell = row[i]             # column H
                cell.alignment = Alignment(horizontal='center')
                cell.font = font

def fill_data(reg_notes):
    not_to_reg = ['Name','Original']
    for num,data in reg_notes.items():
        if len(data['notes']) == 0:
            continue

        create_folder = False
        cont = 0
        for note in data['notes']:
            cont += 1
            for title in TITLES:
                if not title in data: data[title] = ''
                if not title in not_to_reg and note[title] != '' and note['main']:
                    data[title] = note[title]

        if cont > 1:
            create_folder = True

        data['create_folder'] = create_folder
                

def send_messages(reg_notes):
    for num,data in reg_notes.items():
        if 'sent' in data:
            if data['sent']:
                continue
        else:
            data['sent'] = False
        
        if len(data['notes']) == 0 or not data['Dept'] or not data['archived']:
            continue
        
         
        deps = [dp.lower().strip() for dp in data['Dept'].split(',')]

        
        #message = f"Assigned to: *{data['Dept']}* \nLink: {data['link_message']} \nContent: `{data['Content']}`"
        message = f"Content: `{data['Content']}` \nLink: {data['link_message']} \nAssigned to: *{data['Dept']}*"
        
        if data['Ref'] != '':
            message += f"\nRef: _{data['Ref']}_"
        
        if data['Comments'] != '':
            message += f"\nComment: _{data['Comments']}_"
 
        message +=  f"\nRegistry date: {data['Date']}"

        cont = 0
        for dep in deps:
            cont += con.nas.send_message(dep,message)
        
        if len(deps) == cont:
            data['sent'] = True
            logging.info(f"{num} sent to {data['Dept']}")
        else:
            data['sent'] = False

def init_register_despacho():
    logging.info('Starting register despacho -------------------------------')
    reg_notes = read_register_files()
    if reg_notes != {}:
        try:
            fill_data(reg_notes)
        
            archive_notes(reg_notes)
            send_messages(reg_notes)
        except Exception as err:
            raise
            logging.error(err)
            logging.error("Some error archiving and sending notes")
    
        wb_reg = Workbook()
        ws_reg = wb_reg.active
        create_register(ws_reg,reg_notes)
        
        date = datetime.today().strftime('%Y-%m-%d-%HH-%Mm')
        name = f"register-{date}.xlsx"
        upload_register(wb_reg,name,f"{CONFIG['despacho']}/Outbox Despacho")

        
    logging.info('Finishing register despacho ~~~~~~~~~~~~~~~~~~~~~~~~~~~~')

def main():
    PASS = getpass()
    init_register_despacho(PASS)
    
    input("Pulse Enter to continue")



if __name__ == '__main__':
    main()
