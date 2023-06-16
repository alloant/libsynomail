#!/bin/python
import logging
from attrdict import AttrDict

import libsynomail.connection as con
from openpyxl import load_workbook

class Register(AttrDict):
    def __init__(self,flow):
        self.wb = load_workbook(con.nas.download_file(f"{con.CONFIG['folders']['archive']}/Mail {flow} Registry.osheet"))
        self.cg = list(self.wb[f'cg {flow} (1-249)'].iter_rows(values_only=True))
        self.asr = list(self.wb[f'asr {flow} (250-999)'].iter_rows(values_only=True))
        self.ctr = list(self.wb[f'ctr {flow} (from 1000 to 1999)'].iter_rows(values_only=True))
        self.r = list(self.wb[f'r {flow} (2000 onwards)'].iter_rows(values_only=True))

    def get_type(self,no):
        if no < 250:
            tp = 'cg'
        elif no < 1000:
            tp = 'asr'
        elif no < 2000:
            tp = 'ctr'
        else:
            tp = 'r'

        return tp

    def scrap_destination(self,no):
        col = 1 if self.get_type(no) == 'cg' else 0

        for reg in self[self.get_type(no)]:
            if reg[col] == no: break
        
        if self.get_type(no) in ['ctr','r']:
            return reg[2],reg[4]
        else:
            if self.get_type(no) == 'cg':
                return self.get_type(no),reg[6]
            else:
                return self.get_type(no),reg[3]


 
