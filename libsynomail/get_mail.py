#!/bin/python

from datetime import datetime
from pathlib import Path
import logging
import webbrowser

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

from libsynomail import EXT
from libsynomail.classes import File, Note
from libsynomail.scrap_register import Register
from libsynomail.syneml import write_eml

import libsynomail.connection as con

FONT = Font(name= 'Arial',
                size=12,
                bold=False,
                italic=False,
                strike=False,
                underline='none'
                #color='4472C4'
                )

FONT_BOLD = Font(name= 'Arial',
                size=12,
                bold=True,
                italic=False,
                strike=False,
                underline='none'
                #color='4472C4'
                )


def get_notes_in_folders(folders,ctrs):
    team_folders = con.nas.get_team_folders()

    files = []
    paths = []
    for key,path in folders.items():
        if key in ['ctr','dr']:
            for ctr in ctrs:
                value = ctr if 'name' not in ctrs[ctr] else ctrs[ctr]['name']
                paths.append([key,ctr,path.replace('@',value)])
        else:
            paths.append([key,'',path])
    
    for path in paths:
        #if con.CONFIG['DEBUG'] and not (path[1] == 'gul' or path[1] == 'vind1'): continue

        logging.debug(f"Checking path {path}")
        files_in_folder = con.nas.get_file_list(path[2])
        
        if not files_in_folder: continue

        for file in files_in_folder:
            logging.info(f"Found in {path[0]} {path[1]}: {file['name']}")
            source = path[1] if path[0] != 'r' else file['name']
            files.append({"type": path[0],"source": path[1],"file": File(file)})
    
    
    return files

def notes_from_files(files,flow = 'in'):
    notes = {}
    
    files.sort(reverse=True,key = lambda file: f"{file['main']}_{file['file']}")

    for file in files:
        if file['num'] != "":
            note = Note(file['type'],file['source'],file['num'],flow)
            if not note.key in notes:
                notes[note.key] = note
            
            notes[note.key].addFile(file['file'])
        else:
            file.move(f"{con.CONFIG['folders']['despacho']}/Inbox Despacho")
            #con.nas.move(file['file']['file_id'],f"{con.CONFIG['folders']['despacho']}/Inbox Despacho")
            ext = Path(file['file']['name']).suffix[1:]
            if ext in EXT:
                file.convert()
                #con.nas.convert_office(file['file']['file_id'])

    return notes

def generate_register(path_register,notes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Notes"
    ws.append(['Type','Source','No','Year','Ref','Date','Content','Dept','of_annex','Comments','Archived','Sent to'])
    for cell in ws[ws.max_row]:
        cell.alignment = Alignment(horizontal='center')
        cell.font = FONT_BOLD
    
    ws_data = wb.create_sheet(title="Data")
    ws_data.append(['Key','Folder_id','Permanent_link'])
 
    ws_files = wb.create_sheet(title="Files")
    ws_files.append(['Key','Name','Type','Display_data','File_id','Permanent_link','Original_id'])

    
    for key,note in notes.items():
        ws.append(note.exportExcel())
        ws[ws.max_row][5].number_format = 'dd/mm/yyyy;@'
       
        ws_data.append([note.key,note.folder_path,note.permanent_link])

        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal='center')
            cell.font = FONT

        for file in note.files:
            ws_files.append([key] + file.exportExcel())

    column_widths = [10,10,10,12,12,15,50,20,12,50,10,20]
    for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
        ws.column_dimensions[get_column_letter(i)].width = column_width
    
    date = datetime.today().strftime('%Y-%m-%d-%HH-%Mm')
    r_path,r_id,r_link = con.nas.upload_convert_wb(wb,f"register-{date}.xlsx",f"{path_register}")
    
    if r_link != '':
        webbrowser.open(f"https://nas.prome.sg:5001/oo/r/{r_link}")


def rename_file(source,file,new_name = ''):
    try:
        ext = Path(file.name).suffix[1:]
        if not new_name:
            new_name = file.name if source in file.name and source != 'r' else f"{source}_{file.name}"
            new_name = new_name.strip()
        else:
            new_name += f".{ext}"
        
        new_name = new_name.replace('&','and')
        file.rename(new_name)
    except Exception as err:
        logging.error(err)
        logging.error(f"Problem changing name to {file.name}")

def manage_files_despacho(path_files,files,is_from_dr = False):
    flow = 'out' if is_from_dr else 'in'
    notes = notes_from_files(files,flow)
    
    if not notes_from_files: return None
    
    if is_from_dr: #We only need this for the dr to know where they are sending the note
        register = Register('out')
        
    try:
        for note in notes.values():
            for i,file in enumerate(note.files):
                dest = f"{path_files}"
                
                # Getting information about the note from Mail out
                if is_from_dr:
                    note.dept = register.scrap_destination(note.no)
                
                # Getting the key only in first file and changing names
                if i == 0: #It is the main file
                    main_old_name = Path(file.name).stem
                    num = f"0000{note.no}"[-4:]
                    main_name = f"{note.key.split('_')[0]}_{num}"
                    rename_file(note.source,file,main_name)
                else:
                    rename_file(note.source,file,Path(file.name.replace(main_old_name,main_name)).stem)

                # Creating a folder if there are more than 1 file
                if note.of_annex != '':
                    note.folder_id,note.permanent_link = con.nas.create_folder(dest,main_name)
                    dest += f"/{main_name}"
                    note.folder_path = dest
                
                # Moving the note to dest
                file.move(dest)
                
                # Convert the file if needed
                ext = Path(file.name).suffix[1:]
                if ext in EXT:
                    file.convert()

    except Exception as err:
        logging.error(err)
        logging.error("There was some error managing the notes")

    generate_register(f"{path_files}",notes)
    
    return notes

def read_register_file(path_despacho,flow = 'in'):
    files_in_outbox = con.nas.get_file_list(f"{path_despacho}")
    files_in_outbox.sort(reverse = True,key = lambda file: file['name'])
    notes = ''
    for file in files_in_outbox:
        if file['name'][:9] in ["register-"] and 'osheet' in file['name']:
            notes = register_to_notes(file['display_path'],flow)
            break

    return notes


def register_to_notes(register,flow = 'in'):
    wb = load_workbook(con.nas.download_file(register))

    notes_data = list(wb['Notes'].iter_rows(values_only=True))[1:]
    data = list(wb['Data'].iter_rows(values_only=True))[1:]
    files = list(wb['Files'].iter_rows(values_only=True))[1:]
    
    notes = {}
    
    for i,row in enumerate(notes_data):
        no = row[2].replace(' ','').split('","')[1][:-2]
        note = Note(row[0],row[1],no,flow=flow)
        note.year = row[3] if row[3] else ''
        note.ref = row[4] if row[4] else ''
        note.date = row[5] if row[5] else '' 
        note.content = row[6] if row[6] else ''
        note.dept = row[7] if row[7] else '' 
        note.comments = row[9] if row[9] else ''
        note.archived = row[10] if row[10] else ''
        note.sent_to = row[11] if row[11] else ''

        note.folder_path = data[i][1]
        note.permanent_link = data[i][2]

        for file in files:
            if file[0] == note.key:
                note.addFile(File({'name':file[1],'type':file[2],'display_path':file[3],'file_id':file[4],'permanent_link':file[5]},file[6]))

        notes[note.key] = note
    
    return notes


def rec_in_groups(recipients,RECIPIENTS,ctr = True):
    if recipients == 'all':
        for key,rec in RECIPIENTS:
            send_to.append(key)
        
        return list(set(send_to))


    recs = [rec.strip() for rec in recipients.split(',')]
    send_to = []
    ex_groups = []
    in_groups = []

    for rec in recs:
        if rec in RECIPIENTS:
            send_to.append(rec)
        else: #rec is a groups
            if rec.lower() == rec:
                in_groups.append(rec)
            else:
                ex_groups.append(rec)

    for key,rec in RECIPIENTS.items():
        putin = True
        for gp in ex_groups:
            if not gp.lower() in rec['groups']:
                putin = False
                break

        if putin:
            for gp in in_groups:
                if gp in rec['groups']:
                    send_to.append(key)


    return list(set(send_to))

def new_mail_ctr(note):
    send_to = rec_in_groups(note.dept,con.CONFIG['ctrs'],True)
    
    for st in send_to:
        if not st.lower() in note.sent_to.lower():
            for file in note.files:
                file.copy(con.CONFIG['mail_out']['ctr'].replace('@',st))
            note.sent_to += f",{st}" if note.sent_to else st
    
    return True

def new_mail_ctr_sheet(note):
    wb = Workbook()
    ws = wb.active
    ws.append(['No','Date','Content','Ref'])
    ws.append([note.sheetLink(f"cr {note.no}/{note.year[2:]}"),note.date,note.content,note.ref])
    ws[ws.max_row][1].number_format = 'dd/mm/yyyy;@'

    send_to = rec_in_groups(note.dept,con.CONFIG['ctrs'],True)
    
    for st in send_to:
        if not st.lower() in note.sent_to.lower():
            note.sent_to += f",{st}" if note.sent_to else st
            con.nas.upload_convert_wb(wb,f"new_mail.xlsx",con.CONFIG['mail_out']['ctr'].replace('@',st))
    
    return True

def new_mail_eml(note):
    send_to = rec_in_groups(note.dept,con.CONFIG['r'],False)
    TO = []
    for st in send_to:
        if not st.lower() in note.sent_to.lower():
            note.sent_to += f",{st}" if note.sent_to else st
            TO.append(con.CONFIG['r'][st]['email'])
    if TO:
        write_eml(",".join(TO),note,con.CONFIG['folders']['local_folder'])
    return True

def new_mail_asr(note):
    if not 'asr' in note.sent_to.lower():
        for file in note.files:
            file.copy(con.CONFIG['mail_out']['asr'])
        note.sent_to += f",asr" if note.sent_to else 'asr'
    return True

def new_mail_asr_download(note):
    if not 'asr' in note.sent_to:
        for file in note.files:
            con.nas.download_file(file.file_id,f"{CONFIG['folders']['local_folder']}/outbox asr",file.name)
        note.sent_to += f",asr" if note.sent_to else 'asr'
    return True

def register_notes(is_from_dr = False):
    register_dest = con.CONFIG['folders']['to_send'] if is_from_dr else f"{con.CONFIG['folders']['despacho']}/Outbox Despacho"
    flow = 'out' if is_from_dr else 'in'
    notes = read_register_file(register_dest,flow)
    
    if not notes: return None
    
    try:
        for note in notes.values():
            # Moving note to archive if needed
            if not note.archived and note.dept != '':
                if is_from_dr:
                    dest = f"{con.CONFIG['folders']['archive']}/{note.archive_folder}"
                else:
                    dest = f"{con.CONFIG['folders']['archive']}/{note.archive_folder}"
                
                note.move(dest,register_dest)
                note.archived = True

            # Sending copy/message of note to recipient
            if note.archived and note.dept != '':
                if is_from_dr: # Is mail out
                    if note.no < 250: #cg
                        rst = new_mail_eml(note)
                    elif note.no < 1000: #asr
                        rst = new_mail_asr(note)
                    elif note.no < 2000: #ctr
                        rst = new_mail_ctr(note)
                    else: #r
                        rst = new_mail_eml(note)
                else: # Is mail in to one/several dr
                    depts = [dep.lower().strip() for dep in note.dept.split(',')]
                    for dep in depts:
                        if not dep.lower() in note.sent_to.lower():
                            rst = con.nas.send_message(dep,con.CONFIG['deps'],note.message)
                            if rst: note.sent_to += f",{dep}" if note.sent_to else dep
    
    except Exception as err:
        raise
        logging.error(err)
        logging.error("There was some error registering the notes")
        
    generate_register(register_dest,notes)
